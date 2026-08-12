import streamlit as st
import google.generativeai as genai
import PyPDF2
import json
import re
import io
import time
import openpyxl
from datetime import datetime, timedelta

st.set_page_config(page_title="채권자목록 자동 완성기", page_icon="⚖️", layout="centered")

st.title("⚖️ 개인회생 채권자목록 자동 완성 AI")
st.markdown("빈 엑셀 양식과 부채증명서들을 올리면, 법원 제출용 6줄 양식에 맞춰 자동 작성됩니다.")

with st.sidebar:
    st.header("⚙️ 설정")
    api_key = st.text_input("Google Gemini API 키", type="password")
    st.markdown("[API 키 무료 발급받기](https://aistudio.google.com/app/apikey)")

# 파일 업로드 섹션
st.subheader("1. 파일 업로드")
template_file = st.file_uploader("빈 양식 엑셀 파일 (변제계획안.xlsx) 업로드", type="xlsx")
pdf_files = st.file_uploader("부채증명서 PDF 파일 업로드 (여러 개 가능)", type="pdf", accept_multiple_files=True)

if st.button("✨ 자동 작성 시작", type="primary"):
    if not api_key:
        st.error("좌측 메뉴에 API 키를 입력해 주세요.")
    elif not template_file or not pdf_files:
        st.error("엑셀 양식과 부채증명서 PDF를 모두 업로드해 주세요.")
    else:
        # AI 준비
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()

        total_principal = 0
        total_interest = 0

        # PDF 분석
        for i, pdf in enumerate(pdf_files):
            status_text.text(f"[{pdf.name}] 문서에서 데이터 추출 중... 🧐")
            
            # PDF 텍스트 추출 및 개인정보 마스킹 처리
            reader = PyPDF2.PdfReader(pdf)
            text = "".join([page.extract_text() + "\n" for page in reader.pages])
            text = re.sub(r'\d{6}-\d{7}', '******-*******', text) # 주민번호 가림처리
            
            prompt = f"""
            당신은 개인회생 채권자 목록 작성 전문가입니다. 부채증명서에서 다음 정보를 찾아 JSON으로 답하세요.
            [추출 항목]
            - creditor_name: 채권자명 (예: 현대카드(주))
            - start_date: 최초대출일 또는 카드발급일 (예: 2022-08-26, 모르면 "")
            - address: 채권자 주소 (모르면 "")
            - phone: 채권자 대표전화번호 (모르면 "")
            - cause: 채권의 내용 (예: 발급받은신용카드사용채무, 신용대출, 개인대여금 등)
            - principal: 원금 잔액 (숫자만)
            - interest: 이자 잔액 (숫자만, 없으면 0)
            - ref_date: 채무액 산정기준일 (예: 2025-10-28)

            텍스트: {text[:4000]}
            """
            try:
                response = model.generate_content(prompt)
                ai_data = json.loads(response.text.replace('```json', '').replace('```', '').strip())
                
                results.append(ai_data)
                total_principal += int(ai_data['principal'])
                total_interest += int(ai_data['interest'])
            except Exception as e:
                st.warning(f"{pdf.name} 처리 중 오류 발생. 표 구조가 복잡하여 누락되었을 수 있습니다.")
            
            progress_bar.progress((i + 1) / len(pdf_files))
            time.sleep(2)

        status_text.text("✅ AI 분석 완료! 엑셀 양식에 입력합니다...")

        # 엑셀 제어 (openpyxl)
        try:
            wb = openpyxl.load_workbook(template_file)
            if '채권' in wb.sheetnames:
                ws = wb['채권']
            else:
                ws = wb.active # '채권' 탭이 없으면 첫 번째 탭 사용

            # 총합계 입력 (양식의 상단, 원금과 이자 합계 셀)
            # ※ 주의: 실제 엑셀의 셀 위치(D4, D5 등)에 맞게 조정해야 합니다.
            # 샘플 구조를 참고하여 임의로 D4, D5에 넣는 예시입니다.
            ws['D4'] = total_principal 
            ws['D5'] = total_interest

            start_row = 14 # 채권자 1번이 시작되는 행 번호
            
            for idx, data in enumerate(results):
                current_row = start_row + (idx * 6)
                
                # 6줄 규칙에 맞춰 데이터 입력 (열 번호는 A=1, B=2, C=3, D=4...)
                ws.cell(row=current_row, column=1, value=idx + 1) # 채권번호
                ws.cell(row=current_row, column=2, value=data['creditor_name']) # 채권자명
                ws.cell(row=current_row, column=4, value=data['start_date']) # 대출일자
                ws.cell(row=current_row, column=12, value=data['address']) # 주소
                
                ws.cell(row=current_row+1, column=4, value=data['cause']) # 채권의 내용
                ws.cell(row=current_row+1, column=12, value=data['phone']) # 전화번호
                
                ws.cell(row=current_row+2, column=4, value="원금 잔액")
                ws.cell(row=current_row+2, column=6, value=int(data['principal'])) # 실제 원금
                
                # 기준일 다음날 계산
                try:
                    ref_dt = datetime.strptime(data['ref_date'].replace('. ','-').replace('.','-'), "%Y-%m-%d")
                    next_dt = ref_dt + timedelta(days=1)
                    ws.cell(row=current_row+3, column=4, value=next_dt.strftime("%Y-%m-%d"))
                except:
                    ws.cell(row=current_row+3, column=4, value=data['ref_date'])
                
                ws.cell(row=current_row+4, column=4, value=int(data['principal'])) # 원금 다시 기재
                ws.cell(row=current_row+4, column=8, value=f"부채확인서 참조 (산정기준일: {data['ref_date']})")
                
                ws.cell(row=current_row+5, column=4, value=int(data['interest'])) # 이자 기재
                ws.cell(row=current_row+5, column=8, value=f"부채확인서 참조 (산정기준일: {data['ref_date']})")

            # 저장 후 다운로드 버튼 생성
            output = io.BytesIO()
            wb.save(output)
            processed_data = output.getvalue()
            
            st.success("🎉 작성이 완료되었습니다! 아래 버튼을 눌러 결과물을 확인하세요.")
            st.download_button(
                label="📥 완성된 채권자목록 다운로드",
                data=processed_data,
                file_name="완성본_변제계획안.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"엑셀 파일 작성 중 오류가 발생했습니다: {e}")
